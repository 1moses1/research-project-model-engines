"""
Excel Text Extraction Service
"""

import openpyxl


class ExcelExtractor:
    """Extracts text from Excel files"""

    def extract(self, file_path: str) -> str:
        """
        Extract text from Excel file

        Args:
            file_path: Path to Excel file

        Returns:
            Extracted text content
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_content = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                text_content.append(f"\n=== Sheet: {sheet_name} ===\n")

                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None and str(cell).strip():
                            row_text.append(str(cell))

                    if row_text:
                        text_content.append(" | ".join(row_text))

            print(f"📄 Extracted from {len(workbook.sheetnames)} sheets")

            full_text = "\n".join(text_content)

            return full_text

        except Exception as e:
            print(f"⚠️ Excel extraction error: {str(e)}")
            return f"Error extracting Excel: {str(e)}"
